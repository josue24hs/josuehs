from django.http import HttpResponse
from openpyxl import Workbook
from django.shortcuts import get_object_or_404
from django.shortcuts import render, redirect
from .forms import login
from .models import Persona

def agregar(request):
    if request.method == 'POST':
        form = login(request.POST)
        if form.is_valid():
            form.save()
            return redirect('agregar')
    else:
        form = login()
        personas = Persona.objects.all()

    return render(request, 'index.html', {'form': form, 'personas': personas})

def mostrar(request):
    personas = Persona.objects.all()
    return render(request, 'tabla.html', {'personas': personas})

def exportar(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Personas'

    ws['A1'] = 'nombre'
    ws['b1'] = 'edad'
    ws['c1'] = 'diagnostico'
    personas = Persona.objects.all()
    row = 2

    for persona in personas:
        ws[f'A{row}'] = persona.nombre
        ws[f'b{row}'] = persona.edad
        ws[f'c{row}'] = persona.diagnostico
        row +=1
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=personas.xlsx'

    wb.save(response)

    return response

def buscar(request):
    criterio = request.GET.get('opciones','')
    valor = request.GET.get('valor','')
    if criterio and valor:
        if criterio == 'nombre':
            personas = Persona.objects.filter(nombre__icontains=valor)
        elif criterio == 'edad':
            personas = Persona.objects.filter(edad=valor)
        elif criterio == 'diagnostico':
            personas = Persona.objects.filter(diagnostico__icontains=valor)
        elif criterio == 'updrs':
            personas = Persona.objects.filter(updrs=valor)
        else:
            personas = Persona.objects.all()
    else:
        personas = Persona.objects.all()
    return render(request,'tabla.html',{'personas':personas})

def eliminar(request, id):
    persona = get_object_or_404(Persona, id=id)
    persona.delete()
    return redirect('/mostrar/')

def editar(request, id):
    persona = get_object_or_404(Persona, id=id)

    if request.method == 'POST':
        form = login(request.POST, instance=persona)
        if form.is_valid():
            form.save()
            return redirect('mostrar')
    else:
        form = login(instance=persona)

    return render(request, 'editar.html', {'form': form})
